"""PDF を Excel (.xlsx) に変換するユーティリティ。"""
from __future__ import annotations

import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def convert_pdf_to_excel(pdf_path: Path, xlsx_path: Path | None = None) -> Path:
    if xlsx_path is None:
        xlsx_path = pdf_path.with_suffix(".xlsx")

    wb = Workbook()
    ws_text = wb.active
    ws_text.title = "テキスト"
    ws_tables = wb.create_sheet("テーブル一覧")

    text_lines: list[str] = []
    all_tables: list[tuple[int, int, list[list[str]]]] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            page_text = page.extract_text() or ""
            if page_text:
                text_lines.append(f"=== ページ {page_no} ===")
                text_lines.extend(page_text.splitlines())
                text_lines.append("")

            tables = page.extract_tables() or []
            for table_no, table in enumerate(tables, start=1):
                cleaned: list[list[str]] = []
                for row in table:
                    cleaned_row = [clean_cell(cell) for cell in row]
                    if any(cell.strip() for cell in cleaned_row):
                        cleaned.append(cleaned_row)
                if cleaned:
                    all_tables.append((page_no, table_no, cleaned))

    for i, line in enumerate(text_lines, start=1):
        ws_text.cell(row=i, column=1, value=line)
    ws_text.column_dimensions["A"].width = 120

    row_idx = 1
    for page_no, table_no, table in all_tables:
        title = f"ページ{page_no} / テーブル{table_no}"
        cell = ws_tables.cell(row=row_idx, column=1, value=title)
        cell.font = Font(bold=True)
        row_idx += 1

        for row in table:
            for c_idx, value in enumerate(row, start=1):
                cell = ws_tables.cell(row=row_idx, column=c_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1
        row_idx += 1

    for page_no, table_no, table in all_tables:
        sheet_name = f"P{page_no}_T{table_no}"[:31]
        ws = wb.create_sheet(sheet_name)
        max_cols = max(len(row) for row in table)
        for r_idx, row in enumerate(table, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, max_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 24

    wb.save(xlsx_path)
    return xlsx_path


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python pdf_to_excel.py <input.pdf> [output.xlsx]")
        return 1

    pdf_path = Path(sys.argv[1])
    if not pdf_path.exists():
        print(f"File not found: {pdf_path}")
        return 1

    xlsx_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    out = convert_pdf_to_excel(pdf_path, xlsx_path)
    print(f"saved: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
